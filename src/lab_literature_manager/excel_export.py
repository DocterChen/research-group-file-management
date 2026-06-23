"""Excel advanced format export for research outputs."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from .models import Member, Project, ResearchOutput
    from .repository import ResearchRepository


def export_to_excel(repository: ResearchRepository, output_path: Path | str) -> None:
    """导出所有数据到格式化的Excel文件。

    Args:
        repository: 数据仓库实例
        output_path: 输出Excel文件路径

    Raises:
        ImportError: 如果openpyxl未安装
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # 创建各个工作表
    _create_summary_sheet(wb, repository)
    _create_outputs_sheet(wb, repository)
    _create_members_sheet(wb, repository)
    _create_projects_sheet(wb, repository)

    wb.save(output_path)


def _create_summary_sheet(wb: Workbook, repository: ResearchRepository) -> None:
    """创建汇总统计工作表。"""
    ws = wb.create_sheet("汇总统计", 0)

    # 标题
    ws["A1"] = "课题组科研成果统计报告"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="1F4788")
    ws.merge_cells("A1:D1")

    # 基础统计
    summary = repository.build_summary()
    members = repository.list_members()
    projects = repository.list_projects()

    row = 3
    stats = [
        ("总成果数", summary["total_outputs"]),
        ("成员数", len(members)),
        ("项目数", len(projects)),
    ]

    ws["A3"] = "基础统计"
    ws["A3"].font = Font(name="微软雅黑", size=12, bold=True)

    for label, value in stats:
        row += 1
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = Font(name="微软雅黑", size=11)
        ws[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True)

    # 按类型统计
    row += 2
    ws[f"A{row}"] = "成果类型分布"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=12, bold=True)

    type_counts = summary["by_type"]
    for output_type, count in sorted(type_counts.items()):
        row += 1
        ws[f"A{row}"] = output_type
        ws[f"B{row}"] = count

    # 按状态统计
    row += 2
    ws[f"A{row}"] = "审核状态分布"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=12, bold=True)

    status_counts = summary["by_review_status"]
    for status, count in sorted(status_counts.items()):
        row += 1
        ws[f"A{row}"] = status
        ws[f"B{row}"] = count

    # 设置列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def _create_outputs_sheet(wb: Workbook, repository: ResearchRepository) -> None:
    """创建成果清单工作表。"""
    ws = wb.create_sheet("成果清单")

    # 表头
    headers = [
        "成果编号",
        "标题",
        "类型",
        "年份",
        "负责人",
        "参与人",
        "关联项目",
        "审核状态",
        "文章类型",
        "期刊",
        "DOI",
        "投稿状态",
        "创建时间",
        "更新时间",
    ]

    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    outputs = repository.list_outputs()
    for row_idx, output in enumerate(outputs, start=2):
        ws.cell(row=row_idx, column=1, value=output.output_id)
        ws.cell(row=row_idx, column=2, value=output.title)
        ws.cell(row=row_idx, column=3, value=output.output_type.value)
        ws.cell(row=row_idx, column=4, value=output.year)
        ws.cell(row=row_idx, column=5, value=", ".join(output.owner_member_ids))
        ws.cell(row=row_idx, column=6, value=", ".join(output.participant_member_ids))
        ws.cell(row=row_idx, column=7, value=", ".join(output.project_ids))
        ws.cell(row=row_idx, column=8, value=output.review_status.value)

        # 文章专属字段
        if output.article:
            ws.cell(row=row_idx, column=9, value=output.article.article_type)
            ws.cell(row=row_idx, column=10, value=output.article.journal or "")
            ws.cell(row=row_idx, column=11, value=output.article.doi or "")
            ws.cell(row=row_idx, column=12, value=output.article.submission_status or "")

        ws.cell(row=row_idx, column=13, value=output.created_at)
        ws.cell(row=row_idx, column=14, value=output.updated_at)

    # 设置列宽
    column_widths = [15, 40, 12, 8, 20, 20, 20, 12, 12, 25, 20, 12, 20, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 添加筛选器
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(outputs) + 1}"

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置边框
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=1, max_row=len(outputs) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _create_members_sheet(wb: Workbook, repository: ResearchRepository) -> None:
    """创建成员清单工作表。"""
    ws = wb.create_sheet("成员清单")

    # 表头
    headers = ["成员编号", "姓名", "角色", "邮箱", "备注"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    members = repository.list_members()
    for row_idx, member in enumerate(members, start=2):
        ws.cell(row=row_idx, column=1, value=member.member_id)
        ws.cell(row=row_idx, column=2, value=member.name)
        ws.cell(row=row_idx, column=3, value=member.role.value)
        ws.cell(row=row_idx, column=4, value=member.email or "")
        ws.cell(row=row_idx, column=5, value=member.notes or "")

    # 设置列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40

    # 冻结首行
    ws.freeze_panes = "A2"


def _create_projects_sheet(wb: Workbook, repository: ResearchRepository) -> None:
    """创建项目清单工作表。"""
    ws = wb.create_sheet("项目清单")

    # 表头
    headers = ["项目编号", "项目名称", "类型", "负责人", "资助来源", "起始年份", "结束年份"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C28A1A", end_color="C28A1A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    projects = repository.list_projects()
    for row_idx, project in enumerate(projects, start=2):
        ws.cell(row=row_idx, column=1, value=project.project_id)
        ws.cell(row=row_idx, column=2, value=project.name)
        ws.cell(row=row_idx, column=3, value=project.project_type)
        ws.cell(row=row_idx, column=4, value=", ".join(project.owner_member_ids))
        ws.cell(row=row_idx, column=5, value=project.funding_source or "")
        ws.cell(row=row_idx, column=6, value=project.start_year if project.start_year else "")
        ws.cell(row=row_idx, column=7, value=project.end_year if project.end_year else "")

    # 设置列宽
    column_widths = [15, 40, 15, 20, 25, 12, 12]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"
